from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def color_por_valor(valor):
    if valor <= 5:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # rojo
    elif valor <= 7:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # amarillo
    elif valor <= 10:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
    return None

def auto_ajustar_columnas(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2