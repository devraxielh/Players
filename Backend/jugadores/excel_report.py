from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Jugador

def color_por_valor(valor):
    if valor <= 5:
        return PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif valor <= 7:
        return PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    elif valor <= 10:
        return PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    return None

def auto_ajustar_columnas(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = length + 2

def generar_excel_jugador(jugador_id):
    jugador = Jugador.objects.get(pk=jugador_id)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Datos Generales"

    # --- Hoja 1: Datos del jugador
    ws1.merge_cells('A1:E1')
    ws1['A1'] = f"Informe del Jugador: {jugador.nombre_completo}"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')

    datos = [
        ("Nombre completo", jugador.nombre_completo),
        ("Fecha de nacimiento", jugador.fecha_nacimiento.strftime("%d/%m/%Y")),
        ("Posición", jugador.get_posicion_display()),
        ("Estatura (m)", jugador.estatura),
        ("Peso (kg)", jugador.peso),
        ("Observaciones", jugador.observaciones or ""),
    ]
    row = 3
    for label, value in datos:
        ws1[f"A{row}"] = label
        ws1[f"B{row}"] = value
        ws1[f"A{row}"].font = Font(bold=True)
        row += 1

    # --- Estadísticas de Juego
    ws2 = wb.create_sheet("Estadísticas de Juego")
    headers = [
        "Fecha", "Pase Corto", "Pase Largo", "Efectivos", "Errados", "Control",
        "Conducciones", "Remate", "Regate", "Juego Aéreo", "Visión", "Concentración",
        "Ritmo", "Duelos Ofensivos", "Duelos Defensivos"
    ]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header).font = Font(bold=True)

    datos_juego = jugador.estadisticasjuego_set.all().order_by('-fecha')[:5]
    for i, e in enumerate(datos_juego, start=2):
        valores = [
            e.fecha.strftime("%d/%m/%Y"), e.pase_corto, e.pase_largo, e.pase_efectivo,
            e.pase_errado, e.control, e.conducciones, e.remate, e.regate,
            e.juego_aereo, e.vision, e.concentracion, e.ritmo_intensidad,
            e.duelos_ofensivos, e.duelos_defensivos,
        ]
        for col, val in enumerate(valores, 1):
            ws2.cell(row=i, column=col, value=val)

    # --- Promedios juego
    fila_prom = len(datos_juego) + 2
    ws2.cell(row=fila_prom, column=1, value="Promedios").font = Font(bold=True)
    for col in range(2, len(headers) + 1):
        valores = [ws2.cell(row=r, column=col).value for r in range(2, 2 + len(datos_juego)) if isinstance(ws2.cell(row=r, column=col).value, (int, float))]
        if valores:
            promedio = round(sum(valores) / len(valores), 2)
            cell = ws2.cell(row=fila_prom, column=col, value=promedio)
            cell.font = Font(bold=True)
            cell.fill = color_por_valor(promedio)

    # --- Estadísticas Generales
    ws3 = wb.create_sheet("Estadísticas Generales")
    headers_gen = [
        "Fecha", "td_puesto", "td_otros_puestos", "ejecución", "m_defensa", "m_ofensivo",
        "pph", "ppi", "rpi", "rph", "jcabeza", "conduccion", "controles", "finta_regate",
        "agilidad", "velocidad", "resistencia", "fuerza", "flexibilidad",
        "cpsi", "ccog", "crel", "ccomp", "criv", "disci"
    ]
    for col, header in enumerate(headers_gen, 1):
        ws3.cell(row=1, column=col, value=header).font = Font(bold=True)

    datos_gen = jugador.estadisticasgenerales_set.all().order_by('-fecha')[:5]
    for i, e in enumerate(datos_gen, start=2):
        valores = [
            e.fecha, e.td_puesto, e.td_otros_puestos, e.ejecucion, e.m_defensa, e.m_ofensivo,
            e.pph, e.ppi, e.rpi, e.rph, e.jcabeza, e.conduccion, e.controles, e.finta_regate,
            e.agilidad, e.velocidad, e.resistencia, e.fuerza, e.flexibilidad,
            e.cpsi, e.ccog, e.crel, e.ccomp, e.criv, e.disci
        ]
        for col, val in enumerate(valores, 1):
            ws3.cell(row=i, column=col, value=val)

    # --- Promedios generales
    fila_prom_gen = len(datos_gen) + 2
    ws3.cell(row=fila_prom_gen, column=1, value="Promedios").font = Font(bold=True)
    for col in range(2, len(headers_gen) + 1):
        valores = [ws3.cell(row=r, column=col).value for r in range(2, 2 + len(datos_gen)) if isinstance(ws3.cell(row=r, column=col).value, (int, float))]
        if valores:
            promedio = round(sum(valores) / len(valores), 2)
            cell = ws3.cell(row=fila_prom_gen, column=col, value=promedio)
            cell.font = Font(bold=True)
            cell.fill = color_por_valor(promedio)

    auto_ajustar_columnas(ws1)
    auto_ajustar_columnas(ws2)
    auto_ajustar_columnas(ws3)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reporte_{jugador.nombre_completo}.xlsx'
    wb.save(response)
    return response